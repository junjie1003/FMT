import argparse
import numpy as np
import pandas as pd
from scipy.stats import wilcoxon
from itertools import combinations
from openpyxl import load_workbook


def get_args():
    parser = argparse.ArgumentParser(description="Test")
    parser.add_argument("--sheet-name", default="Sheet1", type=str, help="Sheet name")
    parser.add_argument("--table-range", type=str, help="Table range")
    args = parser.parse_args()

    return args


# # NNrepair
# nnrepair = {"intermediate": [87.25, 87.3, 34.61, 81.04, 80.93, 36.3], "last": [87.25, 87, 34.76, 81.04, 80.77, 36.23]}
# dare = [88.5, 88.87, 66.43, 81.3, 81.87, 58.7]

# for layer in ["intermediate", "last"]:
#     print(f"Dare v.s. NNrepair-{layer}")
#     base = nnrepair[layer]
#     w, p = wilcoxon(dare, base, alternative="greater")

#     print("Wilcoxon Test统计量:", w)
#     print("单侧P值:", p)

#     alpha = 0.05

#     if p < alpha:
#         print("拒绝原假设，您的方法的偏差显著小于基线方法的偏差。")
#     else:
#         print("不能拒绝原假设，您的方法的偏差与基线方法的偏差无显著差异。")

args = get_args()

# 读取 Excel 文件
file_path = "dare.xlsx"
wb = load_workbook(file_path)

sheet_name = args.sheet_name
ws = wb[sheet_name]

# 定义各个表格的范围
# table_ranges = ["C2:H11", "I2:N11", "O2:T11", "U2:Z11", "AA2:AF11"]
# table_ranges = ["C34:D43", "E34:F43", "G34:H43", "I34:J43"]
# table_ranges = ["C47:F56", "G47:J56", "K47:N56", "O47:R56", "S47:V56"]
table_ranges = ["C60:D69", "E60:F69", "G60:H69", "I60:J69", "K60:L69"]


# 使用 pandas 读取每个表格并将其存储在字典中
tables = {}
for index, table_range in enumerate(table_ranges):
    data = ws[table_range]
    rows = list(data)

    # 将表格数据转换为列表
    table_data = [[cell.value for cell in row] for row in rows]

    # 转换为 pandas DataFrame
    df = pd.DataFrame(table_data[1:], columns=table_data[0])
    df.replace("-", np.nan, inplace=True)
    df = df.fillna(0)
    tables[index] = df

merged_df = pd.concat([tables[0], tables[1], tables[2], tables[3], tables[4]], axis=0)
# merged_df = pd.concat([tables[0], tables[1], tables[2], tables[3]], axis=0)
merged_df.reset_index(drop=True, inplace=True)
print(merged_df)
column_names = list(df.columns)

# data = ws[args.table_range]
# rows = list(data)

# # 将表格数据转换为列表
# table_data = [[cell.value for cell in row] for row in rows]

# # 转换为 pandas DataFrame
# df = pd.DataFrame(table_data[1:], columns=table_data[0])
# column_names = list(df.columns)
# print(column_names)
# print(df)

baselines = {}
for col in column_names:
    # base = df[col].to_numpy()
    base = merged_df[col].to_numpy()
    baselines[col] = base
print(baselines)

dare = baselines["Dare"]
column_names.remove("Dare")

for col in column_names:
    print(f"Dare v.s. {col}")
    base = baselines[col]
    w, p = wilcoxon(dare, base, alternative="greater")

    print("Wilcoxon Test统计量:", w)
    print("单侧P值:", p)

    alpha = 0.05

    if p < alpha:
        print("拒绝原假设，您的方法的偏差显著小于基线方法的偏差。")
    else:
        print("不能拒绝原假设，您的方法的偏差与基线方法的偏差无显著差异。")
